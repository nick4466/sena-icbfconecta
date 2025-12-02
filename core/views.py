from django.shortcuts import render, redirect, get_object_or_404, reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages
from django.db import transaction
from django.contrib.auth.hashers import make_password
from django.db.models import Q, Count
from .models import Usuario, Rol, Padre, Nino, HogarComunitario, Regional
from django.utils import timezone
from django import forms
from django.contrib.auth.forms import SetPasswordForm
from .forms import AdminPerfilForm, MadrePerfilForm, PadrePerfilForm, NinoForm, PadreForm, CustomAuthForm, AdminForm, NinoSoloForm, BuscarPadreForm, CambiarPadreForm
from desarrollo.models import DesarrolloNino
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import Rol, Usuario, MadreComunitaria, HogarComunitario
from .forms import UsuarioMadreForm, MadreProfileForm, HogarForm 
from django.http import JsonResponse, HttpResponse
from .models import Ciudad
from django.core.paginator import Paginator
from django.template.loader import get_template
import io
from xhtml2pdf import pisa
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from datetime import date, datetime, timedelta
import calendar
from datetime import date
from calendar import monthrange
from django.shortcuts import render
from django.http import JsonResponse
from core.models import Nino
from planeaciones.models import Planeacion
from django.contrib.auth.decorators import login_required
from novedades.models import Novedad
from planeaciones.models import Planeacion
from datetime import datetime as _datetime, date as _date
from core.models import Asistencia
from desarrollo.models import SeguimientoDiario

# --- VISTAS PERSONALIZADAS DE AUTENTICACIÓN ---
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.core.mail import send_mail
from django.conf import settings

def custom_password_reset(request):
    if request.method == 'POST':
        form = PasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            # --- VALIDACIÓN PERSONALIZADA ---
            # Verificar si el correo existe en la base de datos
            associated_users = Usuario.objects.filter(Q(correo__iexact=email))
            if associated_users.exists():
                for user in associated_users:
                    # Lógica de envío de correo (la misma que usa Django por defecto)
                    subject = "Restablecimiento de contraseña para ICBF Conecta"
                    email_template_name = "password_reset/password_reset_email.html"
                    c = {
                        "email": user.email,
                        'domain': request.get_host(),
                        'site_name': 'ICBF Conecta',
                        "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                        "user": user,
                        'token': default_token_generator.make_token(user),
                        'protocol': 'http',
                    }
                    email_content = render_to_string(email_template_name, c)
                    send_mail(subject, email_content, settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=False)
                return redirect('password_reset_done')
            else:
                # Si el correo no existe, añadir un error al formulario
                form.add_error('email', 'No existe una cuenta asociada a este correo electrónico.')
    else:
        form = PasswordResetForm()
    return render(request, 'password_reset/password_reset_form.html', {'form': form})



# --- GENERAR REPORTE PDF DE MATRÍCULA DE UN NIÑO ---
@login_required
def reporte_matricula_nino_pdf(request, nino_id):
    import os
    from django.conf import settings
    
    nino = get_object_or_404(Nino, id=nino_id)
    padre = nino.padre
    hogar = nino.hogar
    usuario_generador = request.user.get_full_name() or request.user.username
    fecha_reporte = timezone.now().strftime('%d/%m/%Y %H:%M')
    
    # Función auxiliar para obtener ruta absoluta de archivo
    def get_absolute_path(file_field):
        if file_field and hasattr(file_field, 'path'):
            try:
                return os.path.abspath(file_field.path)
            except:
                return None
        return None
    
    template = get_template('madre/reporte_ninos.html')
    context = {
        'nino': nino,
        'padre': padre,
        'hogar': hogar,
        'usuario_generador': usuario_generador,
        'fecha_reporte': fecha_reporte,
        # Rutas absolutas para las imágenes del niño
        'nino_foto_path': get_absolute_path(nino.foto),
        'nino_carnet_path': get_absolute_path(nino.carnet_vacunacion),
        'nino_eps_path': get_absolute_path(nino.certificado_eps),
        'nino_registro_path': get_absolute_path(nino.registro_civil_img),
        # Rutas absolutas para las imágenes del padre
        'padre_documento_path': get_absolute_path(padre.documento_identidad_img) if padre else None,
        'padre_sisben_path': get_absolute_path(padre.clasificacion_sisben) if padre else None,
    }
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_matricula_{nino.nombres}_{nino.apellidos}.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def reporte_general_hogar_pdf(request):
    """Genera un reporte PDF con todos los niños del hogar de la madre comunitaria"""
    import os
    from django.conf import settings
    
    # Verificar que el usuario sea madre comunitaria
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Acceso denegado. No tienes los permisos necesarios.')
        return redirect('home')
    
    # Obtener el hogar de la madre comunitaria logueada
    try:
        madre = MadreComunitaria.objects.get(usuario=request.user)
        # El hogar se obtiene a través del related_name 'hogares_asignados'
        hogar = madre.hogares_asignados.first()
        if not hogar:
            messages.error(request, 'No tienes un hogar asignado.')
            return redirect('listar_ninos')
    except MadreComunitaria.DoesNotExist:
        messages.error(request, 'No se encontró información de madre comunitaria.')
        return redirect('listar_ninos')
    
    # Obtener todos los niños del hogar
    ninos = Nino.objects.filter(hogar=hogar).select_related('padre', 'padre__usuario').order_by('apellidos', 'nombres')
    
    if not ninos.exists():
        messages.warning(request, 'No hay niños matriculados en este hogar para generar el reporte.')
        return redirect('listar_ninos')
    
    usuario_generador = request.user.get_full_name() or request.user.username
    fecha_reporte = timezone.now().strftime('%d/%m/%Y %H:%M')
    
    # Función auxiliar para obtener ruta absoluta de archivo
    def get_absolute_path(file_field):
        if file_field and hasattr(file_field, 'path'):
            try:
                return os.path.abspath(file_field.path)
            except:
                return None
        return None
    
    # Preparar datos de cada niño con sus rutas de documentos
    ninos_data = []
    for nino in ninos:
        padre = nino.padre
        nino_info = {
            'nino': nino,
            'padre': padre,
            'nino_foto_path': get_absolute_path(nino.foto),
            'nino_carnet_path': get_absolute_path(nino.carnet_vacunacion),
            'nino_eps_path': get_absolute_path(nino.certificado_eps),
            'nino_registro_path': get_absolute_path(nino.registro_civil_img),
            'padre_documento_path': get_absolute_path(padre.documento_identidad_img) if padre else None,
            'padre_sisben_path': get_absolute_path(padre.clasificacion_sisben) if padre else None,
        }
        ninos_data.append(nino_info)
    
    template = get_template('madre/reporte_general_hogar.html')
    context = {
        'hogar': hogar,
        'ninos_data': ninos_data,
        'total_ninos': ninos.count(),
        'usuario_generador': usuario_generador,
        'fecha_reporte': fecha_reporte,
    }
    
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_general_{hogar.nombre_hogar.replace(" ", "_")}.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)
    return response

@login_required
def certificado_matricula_pdf(request, nino_id):
    """Genera un certificado de matrícula oficial en PDF"""
    import os
    import random
    from datetime import date
    import locale
    
    # Configurar locale para fechas en español
    try:
        locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    except:
        try:
            locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252')
        except:
            pass
    
    nino = get_object_or_404(Nino, id=nino_id)
    padre = nino.padre
    hogar = nino.hogar
    
    # Calcular edad del niño
    hoy = date.today()
    edad = hoy.year - nino.fecha_nacimiento.year - ((hoy.month, hoy.day) < (nino.fecha_nacimiento.month, nino.fecha_nacimiento.day))
    
    # Generar código de verificación único
    codigo_verificacion = f"ICBF-{hogar.id:04d}-{nino.id:05d}-{random.randint(1000, 9999)}"
    
    # Obtener ruta absoluta del logo
    from django.conf import settings
    logo_path = os.path.join(settings.BASE_DIR, 'core', 'static', 'img', 'logo.png')
    if not os.path.exists(logo_path):
        logo_path = None
    else:
        logo_path = os.path.abspath(logo_path)
    
    # Formatear fecha en español manualmente
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    fecha_emision = f"{hoy.day} de {meses[hoy.month - 1]} de {hoy.year}"
    
    # Formatear fechas del niño
    fecha_nac = nino.fecha_nacimiento
    fecha_nacimiento_texto = f"{fecha_nac.day} de {meses[fecha_nac.month - 1]} de {fecha_nac.year}"
    
    fecha_ing = nino.fecha_ingreso
    fecha_ingreso_texto = f"{fecha_ing.day} de {meses[fecha_ing.month - 1]} de {fecha_ing.year}"
    
    template = get_template('madre/certificado_matricula.html')
    context = {
        'nino': nino,
        'padre': padre,
        'hogar': hogar,
        'edad': edad,
        'codigo_verificacion': codigo_verificacion,
        'fecha_emision': fecha_emision,
        'fecha_nacimiento_texto': fecha_nacimiento_texto,
        'fecha_ingreso_texto': fecha_ingreso_texto,
        'año_actual': hoy.year,
        'logo_path': logo_path,
    }
    
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="certificado_matricula_{nino.nombres}_{nino.apellidos}.pdf"'
    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode('utf-8')), dest=response, encoding='utf-8')
    if pisa_status.err:
        return HttpResponse('Error al generar el certificado', status=500)
    return response

from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import Rol, Usuario, MadreComunitaria, HogarComunitario
from .forms import UsuarioMadreForm, MadreProfileForm, HogarForm 
from django.http import JsonResponse
from .models import Ciudad
from django.core.paginator import Paginator

# --- Dependencias para PDF ---
from django.http import HttpResponse
from django.template.loader import get_template
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from xhtml2pdf import pisa
# Asegúrate de importar todos los formularios y modelos necesarios

# ----------------------------------------------------
# 💡 DECORADOR: Restringir acceso por Rol
# ----------------------------------------------------
from functools import wraps

def rol_requerido(nombre_rol):
    """Decorador que verifica si el usuario tiene un rol específico."""
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # Asegurarse de que el usuario esté autenticado y tenga el rol correcto
            if not request.user.is_authenticated or not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != nombre_rol:
                messages.error(request, 'Acceso denegado. No tienes los permisos necesarios.')
                return redirect('home')  # Redirigir a una página de inicio o de error
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

@login_required
def buscar_padre_por_documento(request):
    """
    Vista AJAX para buscar un padre por su documento.
    """
    documento = request.GET.get('documento')
    data = {'encontrado': False}

    if documento:
        usuario_padre = Usuario.objects.filter(documento=documento, rol__nombre_rol='padre').first()
        if usuario_padre:
            padre_profile = Padre.objects.filter(usuario=usuario_padre).first()
            data.update({
                'encontrado': True,
                'nombres': usuario_padre.nombres,
                'apellidos': usuario_padre.apellidos,
                'correo': usuario_padre.correo,
                'telefono': usuario_padre.telefono,
                'direccion': usuario_padre.direccion,
                'ocupacion': padre_profile.ocupacion if padre_profile else ''
            })
    return JsonResponse(data)

# -----------------------------------------------------------------
# 💡 NUEVA FUNCIÓN: Matricular Niño (CRUD Crear)
# -----------------------------------------------------------------

@login_required
def matricular_nino(request):
    # Solo madres comunitarias pueden acceder
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Solo las madres comunitarias pueden matricular niños.')
        return redirect('home')

    # Obtener el hogar de la madre logueada
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')

    if request.method == 'POST':
        padre_form = PadreForm(request.POST, request.FILES, prefix='padre')
        nino_form = NinoForm(request.POST, request.FILES, prefix='nino')
        if padre_form.is_valid() and nino_form.is_valid():
            try:
                with transaction.atomic():
                    rol_padre = Rol.objects.get(nombre_rol='padre')
                    doc_padre = padre_form.cleaned_data['documento']
                    tipo_doc = padre_form.cleaned_data['tipo_documento']
                    
                    # Buscar o crear el usuario del padre
                    usuario_padre, created = Usuario.objects.get_or_create(
                        documento=doc_padre,
                        rol=rol_padre,
                        defaults={
                            'tipo_documento': tipo_doc,
                            'nombres': padre_form.cleaned_data['nombres'],
                            'apellidos': padre_form.cleaned_data['apellidos'],
                            'correo': padre_form.cleaned_data['correo'],
                            'telefono': padre_form.cleaned_data['telefono'],
                            'direccion': padre_form.cleaned_data.get('direccion', ''),
                        }
                    )

                    if created:
                        usuario_padre.set_password(str(doc_padre))
                        usuario_padre.save()
                    else: # Si ya existía, actualizamos sus datos por si hay cambios
                        usuario_padre.tipo_documento = tipo_doc
                        usuario_padre.nombres = padre_form.cleaned_data['nombres']
                        usuario_padre.apellidos = padre_form.cleaned_data['apellidos']
                        usuario_padre.correo = padre_form.cleaned_data['correo']
                        usuario_padre.telefono = padre_form.cleaned_data['telefono']
                        usuario_padre.direccion = padre_form.cleaned_data.get('direccion', '')
                        usuario_padre.save()

                    # Buscar o crear el perfil del padre
                    padre_obj, _ = Padre.objects.get_or_create(usuario=usuario_padre)
                    padre_obj.ocupacion = padre_form.cleaned_data.get('ocupacion', '')
                    padre_obj.otra_ocupacion = padre_form.cleaned_data.get('otra_ocupacion', '')
                    padre_obj.estrato = padre_form.cleaned_data.get('estrato')
                    padre_obj.telefono_contacto_emergencia = padre_form.cleaned_data.get('telefono_contacto_emergencia', '')
                    padre_obj.nombre_contacto_emergencia = padre_form.cleaned_data.get('nombre_contacto_emergencia', '')
                    padre_obj.situacion_economica_hogar = padre_form.cleaned_data.get('situacion_economica_hogar', '')
                    
                    # Guardar archivos del padre si se proporcionan
                    if 'documento_identidad_img' in request.FILES:
                        padre_obj.documento_identidad_img = request.FILES['documento_identidad_img']
                    if 'clasificacion_sisben' in request.FILES:
                        padre_obj.clasificacion_sisben = request.FILES['clasificacion_sisben']
                    
                    padre_obj.save()

                    # Crear el niño y asociarlo
                    nino = nino_form.save(commit=False)
                    nino.hogar = hogar_madre
                    nino.padre = padre_obj
                    # Nuevos campos
                    nino.tipo_sangre = nino_form.cleaned_data.get('tipo_sangre')
                    nino.parentesco = nino_form.cleaned_data.get('parentesco')
                    nino.tiene_discapacidad = nino_form.cleaned_data.get('tiene_discapacidad', False)
                    nino.otra_discapacidad = nino_form.cleaned_data.get('otra_discapacidad', '')
                    nino.otro_pais = nino_form.cleaned_data.get('otro_pais', '')
                    nino.save()
                    # ManyToMany: tipos_discapacidad
                    if nino.tiene_discapacidad:
                        nino.tipos_discapacidad.set(nino_form.cleaned_data.get('tipos_discapacidad'))
                    else:
                        nino.tipos_discapacidad.clear()
                    # Guardar archivos del niño (fotos y documentos)
                    if 'nino-foto' in request.FILES:
                        nino.foto = request.FILES['nino-foto']
                    if 'nino-carnet_vacunacion' in request.FILES:
                        nino.carnet_vacunacion = request.FILES['nino-carnet_vacunacion']
                    if 'nino-certificado_eps' in request.FILES:
                        nino.certificado_eps = request.FILES['nino-certificado_eps']
                    if 'nino-registro_civil_img' in request.FILES:
                        nino.registro_civil_img = request.FILES['nino-registro_civil_img']
                    nino.save()
                    
                    # Guardar el nombre del niño en la sesión para mostrarlo en el SweetAlert
                    request.session['matricula_exitosa'] = {
                        'nombre': f'{nino.nombres} {nino.apellidos}',
                        'mensaje': f'El niño {nino.nombres} {nino.apellidos} ha sido matriculado exitosamente en el hogar {hogar_madre.nombre_hogar}.'
                    }
                    messages.success(request, f'Niño {nino.nombres} matriculado correctamente.')
                    return redirect('listar_ninos')
            except Exception as e:
                messages.error(request, f"Ocurrió un error inesperado: {e}")
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        padre_form = PadreForm(prefix='padre')
        nino_form = NinoForm(prefix='nino')

    # TODO: CAMBIAR TEMPLATE - Para usar el nuevo template mejorado, cambia 'madre/nino_form.html' 
    # por 'madre/nino_form_nuevo.html' (que tiene mejor organización de campos)
    return render(request, 'madre/nino_form_nuevo.html', {
        'hogar_madre': hogar_madre,
        'nino_form': nino_form,
        'padre_form': padre_form,
        'modo_edicion': False
    })
def home(request):
    return render(request, 'home.html')
@login_required
def admin_dashboard(request):
    return render(request, 'admin/dashboard.html')

@login_required
@rol_requerido('administrador')
def admin_reportes(request):
    """Renderiza la página de reportes para el administrador."""
    context = {
        'regionales_filtro': Regional.objects.all().order_by('nombre'),
        'escolaridad_choices': MadreComunitaria.NIVEL_ESCOLARIDAD_CHOICES
    }
    return render(request, 'admin/reportes.html', context)

# ---------- Formularios simples ----------
class MadreForm(forms.ModelForm):
    class Meta:
        model = Usuario
        fields = ['nombres', 'apellidos', 'documento', 'email', 'telefono', 'direccion']


class HogarForm(forms.ModelForm):
    class Meta:
        model = HogarComunitario
        fields = ['nombre_hogar', 'direccion', 'localidad', 'estado', 'regional', 'ciudad']
# En la sección de formularios simples en views.py
class AdministradorForm(forms.ModelForm):
    # Añadir el campo de contraseña, ya que no se incluye automáticamente
    contraseña = forms.CharField(widget=forms.PasswordInput, required=True)
    
    class Meta:
        model = Usuario
        # Elige los campos que quieres:
        fields = ['first_name', 'email', 'rol'] # Asumiendo estos nombres para Usuario

# ---------- CRUD MADRES ----------
@login_required
# En core/views.py

@rol_requerido('administrador')
def listar_madres(request):
    # 💡 MEJORA: Lógica de filtrado
    query_nombre = request.GET.get('nombre', '')
    query_documento = request.GET.get('documento', '')
    query_hogar = request.GET.get('hogar', '')

    # Obtener todos los perfiles de madre con sus datos relacionados
    madres_query = MadreComunitaria.objects.select_related('usuario').prefetch_related('hogares_asignados').all()

    if query_nombre:
        madres_query = madres_query.filter(Q(usuario__nombres__icontains=query_nombre) | Q(usuario__apellidos__icontains=query_nombre))
    if query_documento:
        madres_query = madres_query.filter(usuario__documento__icontains=query_documento)
    if query_hogar:
        madres_query = madres_query.filter(hogares_asignados__nombre_hogar__icontains=query_hogar)
    
    # 💡 CORRECCIÓN: Añadir paginación
    paginator = Paginator(madres_query.distinct(), 5) # 5 madres por página
    page_number = request.GET.get('page')
    madres_paginadas = paginator.get_page(page_number)

    context = {
        'madres': madres_paginadas, # Enviar el objeto paginado a la plantilla
        'filtros': { # Devolver los filtros a la plantilla
            'nombre': query_nombre,
            'documento': query_documento,
            'hogar': query_hogar,
        },
        'paginator': paginator # Opcional, pero útil para la plantilla
    }
    return render(request, 'admin/madres_list.html', context)

@login_required
@rol_requerido('administrador')
def listar_hogares(request):
    # 💡 MEJORA: Lógica de filtrado
    query_nombre = request.GET.get('nombre_hogar', '')
    query_madre = request.GET.get('madre', '')
    query_regional = request.GET.get('regional', '')

    # Consulta base
    hogares = HogarComunitario.objects.select_related(
        'madre__usuario', 'regional'
    ).annotate(
        num_ninos=Count('ninos')
    ).order_by('regional__nombre', 'nombre_hogar')

    if query_nombre:
        hogares = hogares.filter(nombre_hogar__icontains=query_nombre)
    if query_madre:
        hogares = hogares.filter(Q(madre__usuario__nombres__icontains=query_madre) | Q(madre__usuario__apellidos__icontains=query_madre))
    if query_regional:
        hogares = hogares.filter(regional_id=query_regional)

    # 💡 CORRECCIÓN: Añadir paginación
    paginator = Paginator(hogares, 5) # 5 hogares por página
    page_number = request.GET.get('page')
    hogares_paginados = paginator.get_page(page_number)

    context = {
        'hogares': hogares_paginados, # Enviar el objeto paginado
        'regionales_filtro': Regional.objects.all().order_by('nombre'), # Para el dropdown de filtros
        'filtros': {
            'nombre_hogar': query_nombre,
            'madre': query_madre,
            'regional': query_regional,
        },
        'paginator': paginator # Opcional
    }

    return render(request, 'admin/hogares_list.html', context)

# Importa los formularios correctos al inicio de tu views.py
# from .forms import UsuarioMadreForm, MadreProfileForm, HogarForm 
# ...

@login_required
# En core/views.py (con la lógica de importaciones y Rol.objects.get_or_create)


@login_required
@rol_requerido('administrador')
def crear_madre(request):
    rol_madre, _ = Rol.objects.get_or_create(nombre_rol='madre_comunitaria')

    regionales = Regional.objects.all().order_by('nombre')
    if request.method == 'POST':
        usuario_form = UsuarioMadreForm(request.POST, request.FILES)
        madre_profile_form = MadreProfileForm(request.POST, request.FILES)
        hogar_form = HogarForm(request.POST)  # ¡Esta es la línea que faltaba!
        error_step = 1

        if usuario_form.is_valid() and madre_profile_form.is_valid() and hogar_form.is_valid():
            documento = usuario_form.cleaned_data.get('documento')
            # 💡 CORRECCIÓN: Priorizar el nombre del hogar del formulario. Si está vacío, generar uno.
            nombre_hogar = hogar_form.cleaned_data.get('nombre_hogar')
            if not nombre_hogar:
                nombre_hogar = "Hogar de " + usuario_form.cleaned_data.get('nombres', '').split(' ')[0]

            direccion_hogar = hogar_form.cleaned_data.get('direccion')
            localidad = hogar_form.cleaned_data.get('localidad')

            # Validación de documento duplicado
            if Usuario.objects.filter(documento=documento).exists():
                messages.error(request, f"Ya existe un usuario con el documento {documento} registrado.")
                return render(request, 'admin/madres_form.html', {
                    'usuario_form': usuario_form, 'madre_profile_form': madre_profile_form, 'hogar_form': hogar_form,
                    'initial_step': 1, 'regionales': regionales
                })

            # Validación de hogar duplicado
            if HogarComunitario.objects.filter(nombre_hogar__iexact=nombre_hogar, localidad__iexact=localidad).exists():
                messages.error(request, f"Ya existe un hogar llamado '{nombre_hogar}' en la localidad '{localidad}'.")
                error_step = 3
            elif HogarComunitario.objects.filter(direccion__iexact=direccion_hogar).exists():
                messages.error(request, f"La dirección '{direccion_hogar}' ya está registrada para otro hogar.")
                error_step = 3

            if messages.get_messages(request):
                return render(request, 'admin/madres_form.html', {
                    'usuario_form': usuario_form, 'madre_profile_form': madre_profile_form, 'hogar_form': hogar_form,
                    'initial_step': error_step, 'regionales': regionales
                })

            try:
                # Usar transacción atómica para asegurar que todo se crea o nada
                with transaction.atomic():
                    # 1️⃣ Crear usuario
                    usuario = usuario_form.save(commit=False)
                    usuario.rol = rol_madre
                    # La contraseña se establece aquí, puedes cambiarla si es necesario
                    usuario.set_password('123456')
                    usuario.is_active = True
                    usuario.save()

                    # 2️⃣ Crear perfil madre
                    madre_profile = madre_profile_form.save(commit=False)
                    madre_profile.usuario = usuario
                    madre_profile.save()

                    # 3️⃣ Crear hogar comunitario asociado a la madre
                    hogar = hogar_form.save(commit=False)
                    hogar.madre = madre_profile
                    hogar.nombre_hogar = nombre_hogar
                    hogar.save()

                messages.success(request, '¡Madre comunitaria y hogar creados exitosamente! Contraseña por defecto: 123456')
                return redirect('listar_madres')

            except Exception as e:
                messages.error(request, f"Ocurrió un error al guardar los datos: {str(e)}")

        else:
            if usuario_form.errors:
                error_step = 1
            elif madre_profile_form.errors:
                error_step = 2
            elif hogar_form.errors:
                error_step = 3
            messages.error(request, 'Error en los datos suministrados. Revise el paso marcado en azul.')

        return render(request, 'admin/madres_form.html', {
            'usuario_form': usuario_form,
            'madre_profile_form': madre_profile_form,
            'hogar_form': hogar_form,
            'initial_step': error_step, # Para saber en qué paso del formulario mostrar el error
            'regionales': regionales
        })

    # GET
    # 💡 MEJORA: Si el admin tiene una regional, la pre-seleccionamos en el formulario del hogar.
    initial_hogar = {} # Revertido a estado original
    return render(request, 'admin/madres_form.html', {
        'usuario_form': UsuarioMadreForm(),
        'madre_profile_form': MadreProfileForm(),
        # Pasamos los datos iniciales al formulario del hogar
        'hogar_form': HogarForm(),
        'regionales': regionales,
        'initial_step': 1
    })
@login_required
@rol_requerido('administrador')
def editar_madre(request, id):
    # Obtener el usuario que es madre comunitaria
    usuario_madre = get_object_or_404(Usuario, id=id, rol__nombre_rol='madre_comunitaria')
    
    # Obtener el perfil de madre comunitaria
    madre_profile = usuario_madre.madre_profile
    
    # Obtener el hogar comunitario asociado
    hogar = HogarComunitario.objects.filter(madre=madre_profile).first()
    
    if not hogar:
        # Si por alguna razón no tiene hogar, es mejor redirigir.
        messages.error(request, 'Esta madre no tiene un hogar comunitario asignado.')
        return redirect('listar_madres')

    if request.method == 'POST':
        # PASAR request.FILES para permitir subida de nueva foto
        usuario_form = UsuarioMadreForm(request.POST, instance=usuario_madre)
        madre_profile_form = MadreProfileForm(request.POST, request.FILES, instance=madre_profile)
        hogar_form = HogarForm(request.POST, instance=hogar)
        
        if usuario_form.is_valid() and madre_profile_form.is_valid() and hogar_form.is_valid():
            # Validación para hogar duplicado, excluyendo el hogar actual
            nombre_hogar = hogar_form.cleaned_data['nombre_hogar']
            direccion_hogar = hogar_form.cleaned_data['direccion']
            localidad = hogar_form.cleaned_data.get('localidad')

            if HogarComunitario.objects.filter(nombre_hogar__iexact=nombre_hogar, localidad__iexact=localidad).exclude(id=hogar.id).exists():
                messages.error(request, f"Ya existe otro hogar llamado '{nombre_hogar}' en la localidad de '{localidad}'.")
            elif HogarComunitario.objects.filter(direccion__iexact=direccion_hogar).exclude(id=hogar.id).exists():
                messages.error(request, f"La dirección '{direccion_hogar}' ya está registrada para otro hogar.")

            # Si hay mensajes de error, renderizar de nuevo el formulario
            if messages.get_messages(request):
                return render(request, 'admin/madres_form.html', {
                    'usuario_form': usuario_form,
                    'madre_profile_form': madre_profile_form,
                    'hogar_form': hogar_form,
                    'modo_edicion': True,
                    'regionales': Regional.objects.all().order_by('nombre')
                })

            try:
                with transaction.atomic():
                    # Guardar usuario
                    usuario_actualizado = usuario_form.save()
                    
                    # Guardar perfil y hogar
                    madre_profile_form.save()
                    hogar_form.save()
                
                messages.success(request, '¡Madre comunitaria y hogar actualizados exitosamente!')
                return redirect('listar_madres') # Corregido para que el redirect esté dentro del try
            except Exception as e:
                messages.error(request, f'Error al guardar los cambios: {str(e)}')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        usuario_form = UsuarioMadreForm(instance=usuario_madre)
        madre_profile_form = MadreProfileForm(instance=madre_profile)
        hogar_form = HogarForm(instance=hogar)

    regionales = Regional.objects.all().order_by('nombre')
    return render(request, 'admin/madres_form.html', {
        'usuario_form': usuario_form,
        'madre_profile_form': madre_profile_form,
        'hogar_form': hogar_form,
        'modo_edicion': True,
        'regionales': regionales
    })


@login_required
def eliminar_madre(request, id):
    # Obtener el usuario que es madre comunitaria
    usuario_madre = get_object_or_404(Usuario, id=id, rol__nombre_rol='madre_comunitaria')
    
    try:
        # Usar una transacción para asegurar la integridad de los datos
        with transaction.atomic():
            # El perfil de MadreComunitaria se eliminará en cascada cuando se elimine el Usuario,
            # pero el HogarComunitario está protegido.
            # Primero, eliminamos los hogares asociados a su perfil.
            if hasattr(usuario_madre, 'madre_profile'):
                HogarComunitario.objects.filter(madre=usuario_madre.madre_profile).delete()
            # Ahora sí podemos eliminar el usuario (y su perfil de madre en cascada)
            usuario_madre.delete()
        messages.success(request, '¡Madre comunitaria y su hogar asociado han sido eliminados exitosamente!')
    except Exception as e:
        messages.error(request, f"Ocurrió un error al intentar eliminar a la madre: {e}")
        
    return redirect('listar_madres')

@login_required
@rol_requerido('administrador')
def listar_administradores(request):
    # 💡 MEJORA: Lógica de filtrado
    query_nombre = request.GET.get('nombre', '')
    query_documento = request.GET.get('documento', '')

    rol_admin, _ = Rol.objects.get_or_create(nombre_rol='administrador')
    # Incluimos la regional en la consulta para optimizar
    administradores = Usuario.objects.filter(rol=rol_admin).order_by('nombres')

    if query_nombre:
        administradores = administradores.filter(Q(nombres__icontains=query_nombre) | Q(apellidos__icontains=query_nombre))
    if query_documento:
        administradores = administradores.filter(documento__icontains=query_documento)

    # 💡 CORRECCIÓN: Añadir paginación
    paginator = Paginator(administradores, 5) # 5 administradores por página
    page_number = request.GET.get('page')
    admins_paginados = paginator.get_page(page_number)

    context = {
        'administradores': admins_paginados, # Enviar el objeto paginado
        'filtros': {'nombre': query_nombre, 'documento': query_documento},
        'paginator': paginator # Opcional
    }
    return render(request, 'admin/administradores_list.html', context)


@login_required
@rol_requerido('administrador')
def crear_administrador(request):
    rol_admin, _ = Rol.objects.get_or_create(nombre_rol='administrador')
    
    if request.method == 'POST':
        form = AdminForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            correo = form.cleaned_data['correo']
            contraseña = form.cleaned_data['contraseña']

            if not contraseña:
                messages.error(request, 'El campo contraseña es obligatorio para crear un nuevo administrador.')
                return render(request, 'admin/administradores_form.html', {'form': form})

            if Usuario.objects.filter(Q(documento=documento) | Q(correo=correo)).exists():
                messages.error(request, 'Ya existe un usuario con ese documento o correo electrónico.')
                return render(request, 'admin/administradores_form.html', {'form': form})

            usuario = form.save(commit=False)
            usuario.rol = rol_admin
            usuario.set_password(contraseña)
            usuario.save()

            messages.success(request, '¡Administrador creado exitosamente!')
            return redirect('listar_administradores')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        form = AdminForm()

    return render(request, 'admin/administradores_form.html', {'form': form})

@login_required
@rol_requerido('administrador')
def editar_administrador(request, id):
    admin = get_object_or_404(Usuario, id=id, rol__nombre_rol='administrador')
    if request.method == 'POST':
        form = AdminForm(request.POST, request.FILES, instance=admin)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            correo = form.cleaned_data['correo']

            if Usuario.objects.filter(Q(documento=documento) | Q(correo=correo)).exclude(id=id).exists():
                messages.error(request, 'Ya existe otro usuario con ese documento o correo electrónico.')
                return render(request, 'admin/administradores_form.html', {'form': form, 'admin': admin})

            usuario = form.save(commit=False)
            nueva_contraseña = form.cleaned_data.get('contraseña')
            if nueva_contraseña:
                usuario.set_password(nueva_contraseña)
                if request.user.id == admin.id:
                    update_session_auth_hash(request, usuario)
            usuario.save()
            messages.success(request, '¡Administrador actualizado exitosamente!')
            return redirect('listar_administradores')
    else:
        form = AdminForm(instance=admin)

    return render(request, 'admin/administradores_form.html', {'form': form, 'admin': admin})

def _setup_excel_report_header(ws, title, record_count, num_columns):
    """
    Función auxiliar para configurar el encabezado personalizado en los reportes de Excel.
    """
    # --- ESTILOS ---
    title_font = Font(name='Poppins', bold=True, size=16)
    info_bar_font = Font(name='Poppins', bold=True, color='FFFFFF')
    info_bar_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    info_label_font = Font(name='Poppins', bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- TÍTULO PRINCIPAL ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    ws.row_dimensions[1].height = 30

    # --- BARRA DE INFORMACIÓN ---
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_columns)
    info_bar_cell = ws.cell(row=2, column=1, value="Información del reporte")
    info_bar_cell.font = info_bar_font
    info_bar_cell.fill = info_bar_fill
    info_bar_cell.alignment = center_alignment
    ws.row_dimensions[2].height = 25

    # --- Lógica para dividir las columnas de la sección informativa ---
    # Si hay 6 columnas, el punto medio será 3. El label irá de 1 a 3, y el valor de 4 a 6.
    mid_point_col = num_columns // 2

    # --- SECCIÓN INFORMATIVA ---
    # Fecha de Generación
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=mid_point_col)
    fecha_label = ws.cell(row=3, column=1, value="Fecha de Generación:")
    fecha_label.font = info_label_font

    ws.merge_cells(start_row=3, start_column=mid_point_col + 1, end_row=3, end_column=num_columns)
    fecha_value = ws.cell(row=3, column=mid_point_col + 1, value=timezone.now().strftime('%Y-%m-%d %H:%M:%S'))
    fecha_value.alignment = left_alignment

    # Aplicar borde a toda la fila 3
    for col in range(1, num_columns + 1):
        ws.cell(row=3, column=col).border = thin_border

    # Total de Registros
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=mid_point_col)
    total_label = ws.cell(row=4, column=1, value="Total de Registros:")
    total_label.font = info_label_font

    ws.merge_cells(start_row=4, start_column=mid_point_col + 1, end_row=4, end_column=num_columns)
    total_value = ws.cell(row=4, column=mid_point_col + 1, value=record_count)
    total_value.alignment = left_alignment

    # Aplicar borde a toda la fila 4
    for col in range(1, num_columns + 1):
        ws.cell(row=4, column=col).border = thin_border

    # Ajustar ancho de la primera columna
    ws.column_dimensions['A'].width = 25

    # Devolver la fila donde deben empezar los datos de la tabla
    return 5 # Los datos comenzarán en la fila 5

@login_required
@rol_requerido('administrador')
def reporte_administradores_excel(request):
    # Obtener filtros de la URL
    nombre = request.GET.get('nombre', '')
    documento = request.GET.get('documento', '')

    # Filtrar administradores
    rol_admin, _ = Rol.objects.get_or_create(nombre_rol='administrador')
    administradores = Usuario.objects.filter(rol=rol_admin).order_by('nombres')
    if nombre:
        administradores = administradores.filter(Q(nombres__icontains=nombre) | Q(apellidos__icontains=nombre))
    if documento:
        administradores = administradores.filter(documento__icontains=documento)

    # --- Encabezados y configuración inicial ---
    headers = ['Nombres', 'Apellidos', 'Tipo Documento', 'Documento', 'Correo', 'Teléfono']
    num_columns = len(headers)

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Administradores'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte De Administradores", administradores.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25

    # --- Llenar datos ---
    for admin in administradores:
        start_row += 1
        ws.cell(row=start_row, column=1, value=admin.nombres)
        ws.cell(row=start_row, column=2, value=admin.apellidos)
        ws.cell(row=start_row, column=3, value=admin.get_tipo_documento_display())
        ws.cell(row=start_row, column=4, value=admin.documento)
        ws.cell(row=start_row, column=5, value=admin.correo)
        ws.cell(row=start_row, column=6, value=admin.telefono)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Preparar la respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_administradores.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('administrador')
def reporte_madres_excel(request):
    # Obtener filtros
    nombre = request.GET.get('nombre')
    hogar_asignado = request.GET.get('hogar') # Corregido para coincidir con el filtro de la lista
    escolaridad = request.GET.get('escolaridad', None)

    # Filtrar madres
    madres = MadreComunitaria.objects.select_related('usuario').prefetch_related('hogares_asignados').order_by('usuario__nombres')
    if nombre:
        madres = madres.filter(Q(usuario__nombres__icontains=nombre) | Q(usuario__apellidos__icontains=nombre))
    if hogar_asignado and hogar_asignado != '':
        madres = madres.filter(hogares_asignados__nombre_hogar__icontains=hogar_asignado)
    if escolaridad and escolaridad != '':
        madres = madres.filter(nivel_escolaridad=escolaridad)
    
    madres_list = madres.distinct()

    # --- Encabezados y configuración inicial ---
    headers = ['Nombres', 'Apellidos', 'Correo', 'Documento', 'Nivel de Escolaridad', 'Hogar Asignado']
    num_columns = len(headers)

    # --- Crear libro de Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Madres Comunitarias'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte De Madres Comunitarias", madres_list.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25
        
    # --- Llenar datos ---
    for madre in madres_list:
        hogar = madre.hogares_asignados.first()
        start_row += 1
        row_data = [madre.usuario.nombres, madre.usuario.apellidos, madre.usuario.correo, madre.usuario.documento, madre.get_nivel_escolaridad_display(), hogar.nombre_hogar if hogar else 'N/A']
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=start_row, column=col_num, value=cell_value)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_madres_comunitarias.xlsx"'
    wb.save(response)
    return response

@login_required
@rol_requerido('administrador')
def reporte_hogares_excel(request):
    # Obtener filtros
    nombre_hogar = request.GET.get('nombre_hogar')
    regional_id = request.GET.get('regional', '')
    ciudad = request.GET.get('ciudad', '')
    madre_comunitaria = request.GET.get('madre', '') # Corregido para coincidir con el filtro de la lista
    ninos_matriculados = request.GET.get('ninos_matriculados')

    # Filtrar hogares
    hogares = HogarComunitario.objects.select_related('madre__usuario', 'regional', 'ciudad').annotate(
        num_ninos=Count('ninos')
    ).order_by('regional__nombre', 'nombre_hogar')
    if nombre_hogar:
        hogares = hogares.filter(nombre_hogar__icontains=nombre_hogar)
    if regional_id:
        hogares = hogares.filter(regional_id=regional_id)
    if ciudad:
        hogares = hogares.filter(ciudad__nombre__icontains=ciudad)
    if madre_comunitaria:
        hogares = hogares.filter(Q(madre__usuario__nombres__icontains=madre_comunitaria) | Q(madre__usuario__apellidos__icontains=madre_comunitaria))
    if ninos_matriculados:
        try:
            hogares = hogares.filter(num_ninos=int(ninos_matriculados))
        except (ValueError, TypeError):
            pass

    # --- Encabezados y configuración inicial ---
    headers = ['Nombre del Hogar', 'Madre Comunitaria', 'Regional', 'Ciudad', 'Dirección', 'Niños Matriculados', 'Capacidad', 'Estado']
    num_columns = len(headers)

    # --- Crear libro de Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Hogares Comunitarios'

    # --- Generar encabezado del reporte ---
    start_row = _setup_excel_report_header(ws, "Reporte De Hogares Comunitarios", hogares.count(), num_columns)

    # --- ESTILOS PARA LA TABLA DE DATOS ---
    header_font = Font(name='Poppins', bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='004080', end_color='004080', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Escribir encabezados de la tabla de datos ---
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    ws.row_dimensions[start_row].height = 25

    # --- Llenar datos ---
    for hogar in hogares:
        start_row += 1
        row_data = [hogar.nombre_hogar, f"{hogar.madre.usuario.nombres} {hogar.madre.usuario.apellidos}", hogar.regional.nombre if hogar.regional else 'N/A', hogar.ciudad.nombre if hogar.ciudad else 'N/A', hogar.direccion, hogar.num_ninos, hogar.capacidad_maxima, hogar.get_estado_display()]
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=start_row, column=col_num, value=cell_value)

    # --- AJUSTAR ANCHO DE COLUMNAS Y BORDES ---
    for col_num, header_title in enumerate(headers, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for cell in ws[column_letter]:
            cell.border = thin_border
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_hogares_comunitarios.xlsx"'
    wb.save(response)
    return response

@login_required
def eliminar_administrador(request, id):
    Usuario.objects.filter(id=id).delete()
    return redirect('listar_administradores')
# ... Tus otras funciones (home, admin_dashboard, crear_madre, etc.) ...

# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Redirección por Rol (Se ejecuta después del login)
# ----------------------------------------------------
@login_required
def role_redirect(request):
    """
    Redirige al dashboard apropiado según el rol del usuario.
    Esta será la URL de redirección principal después de un login exitoso.
    """
    if not request.user.rol:
        return redirect('home')

    role = request.user.rol.nombre_rol.lower()

    if role == 'administrador':
        return redirect('admin_dashboard')
    elif role == 'madre_comunitaria':
        return redirect('madre_dashboard')
    elif role == 'padre':
        return redirect('padre_dashboard')

    return redirect('home')

# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Dashboard de la Madre Comunitaria
# ----------------------------------------------------
@login_required
def madre_dashboard(request):
    from django.db.models import Count, Q, F, Value, CharField
    from django.utils import timezone
    from datetime import datetime, timedelta
    from planeaciones.models import Planeacion
    from novedades.models import Novedad
    import json
    
    if request.user.rol.nombre_rol != 'madre_comunitaria':
        return redirect('role_redirect')

    # Obtener hogar de la madre
    hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
    
    if not hogar_madre:
        return render(request, 'madre/dashboard.html', {'error': 'No tienes un hogar asignado.'})
    
    # Obtener todos los niños del hogar
    ninos = Nino.objects.filter(hogar=hogar_madre).select_related('padre', 'padre__usuario')
    total_ninos = ninos.count()
    
    # Información del hogar
    capacidad_maxima = hogar_madre.capacidad_maxima
    disponibles = capacidad_maxima - total_ninos
    porcentaje_ocupacion = round((total_ninos / capacidad_maxima) * 100) if capacidad_maxima > 0 else 0
    
    # Datos de género
    ninos_masculino = ninos.filter(genero='masculino').count()
    ninos_femenino = ninos.filter(genero='femenino').count()
    
    # Calcular edades
    hoy = datetime.now().date()
    edades_data = []
    edad_promedio = 0
    edad_minima = None
    edad_maxima = None
    
    for nino in ninos:
        if nino.fecha_nacimiento:
            edad = (hoy - nino.fecha_nacimiento).days // 365
            edades_data.append(edad)
            if edad_minima is None or edad < edad_minima:
                edad_minima = edad
            if edad_maxima is None or edad > edad_maxima:
                edad_maxima = edad
    
    if edades_data:
        edad_promedio = round(sum(edades_data) / len(edades_data), 1)
    
    # Distribución por edad (para gráfica)
    edad_0_2 = sum(1 for e in edades_data if e <= 2)
    edad_3_4 = sum(1 for e in edades_data if 3 <= e <= 4)
    edad_5 = sum(1 for e in edades_data if e >= 5)
    
    # Documentos
    ninos_con_cedula = ninos.filter(documento__isnull=False).count()
    ninos_con_vacunas = ninos.filter(carnet_vacunacion__isnull=False).count()
    ninos_con_afiliacion = ninos.filter(certificado_eps__isnull=False).count()
    
    # Asistencia últimas 4 semanas
    fecha_inicio = hoy - timedelta(days=28)
    asistencias = Asistencia.objects.filter(
        nino__hogar=hogar_madre,
        fecha__gte=fecha_inicio
    )
    
    asistencias_presentes = asistencias.filter(estado='Presente').count()
    asistencias_ausentes = asistencias.filter(estado='Ausente').count()
    total_dias_registro = (total_ninos * 28) if total_ninos > 0 else 1
    porcentaje_asistencia = round((asistencias_presentes / total_dias_registro) * 100) if total_dias_registro > 0 else 0
    
    # Datos de asistencia por niño (para tabla)
    ninos_asistencia = []
    for nino in ninos:
        asistencias_nino = Asistencia.objects.filter(
            nino=nino,
            fecha__gte=fecha_inicio
        )
        faltas = asistencias_nino.filter(estado='Ausente').count()
        asistencias_nino_count = asistencias_nino.filter(estado='Presente').count()
        porcentaje_nino = round((asistencias_nino_count / 28) * 100) if asistencias_nino_count > 0 else 0
        
        # Calcular edad
        edad = 0
        if nino.fecha_nacimiento:
            edad = (hoy - nino.fecha_nacimiento).days // 365
        
        # Determinar estado
        if faltas >= 7:
            estado = 'Alto riesgo'
            icono_estado = 'alert'
        elif faltas >= 4:
            estado = 'Advertencia'
            icono_estado = 'warning'
        else:
            estado = 'Normal'
            icono_estado = 'check'
        
        # Verificar documentos
        documentos_completos = 0
        if nino.documento:
            documentos_completos += 1
        if nino.carnet_vacunacion:
            documentos_completos += 1
        if nino.certificado_eps:
            documentos_completos += 1
        
        ninos_asistencia.append({
            'nino': nino,
            'edad': edad,
            'genero': nino.get_genero_display(),
            'faltas': faltas,
            'asistencias': asistencias_nino_count,
            'porcentaje': porcentaje_nino,
            'estado': estado,
            'icono_estado': icono_estado,
            'documentos': documentos_completos,
            'documentos_total': 3
        })
    
    # Novedades recientes
    fecha_novedades = hoy - timedelta(days=7)
    novedades_recientes = Novedad.objects.filter(
        nino__hogar=hogar_madre,
        fecha__gte=fecha_novedades
    ).select_related('nino').order_by('-fecha')[:3]
    
    # Planeaciones de esta semana
    inicio_semana = hoy - timedelta(days=hoy.weekday())
    planeaciones_semana = Planeacion.objects.filter(
        madre=request.user,
        fecha__gte=inicio_semana,
        fecha__lte=inicio_semana + timedelta(days=6)
    ).count()
    
    # Datos para gráficas
    genero_data = json.dumps({
        'Masculino': ninos_masculino,
        'Femenino': ninos_femenino
    })
    
    edad_distribucion = json.dumps({
        '0-2 años': edad_0_2,
        '3-4 años': edad_3_4,
        '5+ años': edad_5
    })
    
    documentos_data = json.dumps({
        'Cédula': ninos_con_cedula,
        'Vacunas': ninos_con_vacunas,
        'Salud': ninos_con_afiliacion
    })
    
    # Calcular niños con documentos faltantes
    ninos_documentos_faltantes = []
    ninos_sin_documentos_completos = 0
    
    for nino in ninos:
        documentos_faltantes = []
        
        # Verificar registro civil
        if not nino.registro_civil_img:
            documentos_faltantes.append({
                'nombre': 'Registro Civil',
                'campo': 'registro_civil_img',
                'requerido': True
            })
        
        # Verificar carnet de vacunación
        if not nino.carnet_vacunacion:
            documentos_faltantes.append({
                'nombre': 'Carnet de Vacunación',
                'campo': 'carnet_vacunacion',
                'requerido': True
            })
        
        # Verificar certificado EPS
        if not nino.certificado_eps:
            documentos_faltantes.append({
                'nombre': 'Certificado de Afiliación a Salud (EPS)',
                'campo': 'certificado_eps',
                'requerido': True
            })
        
        # Verificar foto
        if not nino.foto:
            documentos_faltantes.append({
                'nombre': 'Foto del Niño',
                'campo': 'foto',
                'requerido': False
            })
        
        if documentos_faltantes:
            ninos_sin_documentos_completos += 1
            ninos_documentos_faltantes.append({
                'id': nino.id,
                'nombre': f"{nino.nombres} {nino.apellidos}",
                'documentos_faltantes': documentos_faltantes
            })
    
    # Convertir a JSON para JavaScript
    ninos_documentos_faltantes_json = json.dumps(ninos_documentos_faltantes)
    
    # Nombre completo de la madre
    nombre_madre = f"{request.user.nombres} {request.user.apellidos}"
    
    # Fecha de creación del hogar
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    if hogar_madre.fecha_registro:
        fecha_registro_hogar = hogar_madre.fecha_registro
        fecha_creacion_hogar = f"{fecha_registro_hogar.day} de {meses[fecha_registro_hogar.month - 1]} de {fecha_registro_hogar.year}"
        hora_creacion_hogar = fecha_registro_hogar.strftime('%H:%M')
    else:
        fecha_creacion_hogar = "No disponible"
        hora_creacion_hogar = "--:--"
    
    context = {
        'nombre_madre': nombre_madre,
        'hogar_madre': hogar_madre,
        'hora_actual': hora_creacion_hogar,
        'fecha_actual': fecha_creacion_hogar,
        # Capacidad
        'total_ninos': total_ninos,
        'capacidad_maxima': capacidad_maxima,
        'disponibles': disponibles,
        'porcentaje_ocupacion': porcentaje_ocupacion,
        # Información de niños
        'ninos': ninos,
        'ninos_asistencia': ninos_asistencia,
        'ninos_masculino': ninos_masculino,
        'ninos_femenino': ninos_femenino,
        # Edades
        'edad_promedio': edad_promedio,
        'edad_minima': edad_minima,
        'edad_maxima': edad_maxima,
        # Documentos
        'ninos_con_cedula': ninos_con_cedula,
        'ninos_con_vacunas': ninos_con_vacunas,
        'ninos_con_afiliacion': ninos_con_afiliacion,
        'ninos_sin_documentos_completos': ninos_sin_documentos_completos,
        'ninos_documentos_faltantes': ninos_documentos_faltantes_json,
        # Asistencia
        'asistencias_presentes': asistencias_presentes,
        'asistencias_ausentes': asistencias_ausentes,
        'porcentaje_asistencia': porcentaje_asistencia,
        # Gráficas
        'genero_data': genero_data,
        'edad_distribucion': edad_distribucion,
        'documentos_data': documentos_data,
        # Otros
        'novedades_recientes': novedades_recientes,
        'planeaciones_semana': planeaciones_semana,
    }

    return render(request, 'madre/dashboard.html', context)


# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Dashboard del Padre de Familia
# ----------------------------------------------------
@login_required
def padre_dashboard(request):
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')
    # Importar modelos necesarios aquí para evitar importaciones circulares
    from novedades.models import Novedad
    from desarrollo.models import DesarrolloNino

    try:
        padre = Padre.objects.get(usuario=request.user)
        ninos_qs = Nino.objects.filter(padre=padre).order_by('nombres')

        ninos_data = []
        for nino in ninos_qs:
            # Obtener la última asistencia registrada para el niño
            ultima_asistencia_obj = Asistencia.objects.filter(nino=nino).order_by('-fecha').first()
            if ultima_asistencia_obj:
                asistencia_info = {
                    'estado': ultima_asistencia_obj.estado,
                    'mensaje': f"El día {ultima_asistencia_obj.fecha.strftime('%d/%m/%Y')} estuvo {ultima_asistencia_obj.estado.lower()}."
                }
            else:
                asistencia_info = None

            # Obtener el último desarrollo y la última novedad
            ultimo_desarrollo = DesarrolloNino.objects.filter(nino=nino).order_by('-fecha_fin_mes').first()
            ultima_novedad = Novedad.objects.filter(nino=nino).order_by('-fecha').first()

            ninos_data.append({
                'nino': nino,
                'ultima_asistencia': asistencia_info,
                'ultimo_desarrollo': ultimo_desarrollo,
                'ultima_novedad': ultima_novedad
            })

        return render(request, 'padre/dashboard.html', {
            'ninos_data': ninos_data,
        })
    except Padre.DoesNotExist:
        # Manejar el caso donde el padre no tiene un hijo asignado
        return render(request, 'padre/dashboard.html', {'error': 'No tienes un niño asignado.'})

def _get_logro_style(logro):
    """
    Retorna un color y un ícono para el logro mensual.
    """
    logro = (logro or "").lower()
    if "excelente" in logro or "alto" in logro:
        return "#27ae60", "fas fa-star"
    elif "medio" in logro or "regular" in logro:
        return "#f1c40f", "fas fa-check"
    elif "bajo" in logro or "alerta" in logro:
        return "#e74c3c", "fas fa-exclamation"
    else:
        return "#9B59B6", "fas fa-child"


# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Ver Desarrollo (Vista del Padre)
# ----------------------------------------------------
@login_required
def padre_ver_desarrollo(request, nino_id):
    if request.user.rol.nombre_rol != 'padre':
        return redirect('role_redirect')

    try:
        padre = Padre.objects.get(usuario=request.user)
        nino = get_object_or_404(Nino, id=nino_id, padre=padre)

        desarrollos_qs = DesarrolloNino.objects.filter(nino=nino).order_by('-fecha_fin_mes')

        mes_filtro = request.GET.get('mes', '')
        if mes_filtro:
            try:
                year, month = map(int, mes_filtro.split('-'))
                desarrollos_qs = desarrollos_qs.filter(fecha_fin_mes__year=year, fecha_fin_mes__month=month)
            except (ValueError, TypeError):
                mes_filtro = ''

        # Procesamiento de datos para la plantilla
        desarrollos_list = []
        for desarrollo in desarrollos_qs:
            accent_color, icono = _get_logro_style(desarrollo.logro_mes)

            desarrollos_list.append({
                'desarrollo': desarrollo,
                'accent_color': accent_color,
                'icono': icono,
            })

        # Paginación
        paginator = Paginator(desarrollos_list, 2) # 2 registros por página
        page_number = request.GET.get('page')
        desarrollos_paginados = paginator.get_page(page_number)

        return render(request, 'padre/desarrollo_list.html', {
            'nino': nino,
            'desarrollos': desarrollos_paginados,
            'filtros': {
                'mes': mes_filtro
            }
        })
    except (Padre.DoesNotExist, Nino.DoesNotExist):
        return redirect('padre_dashboard') # pragma: no cover

MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}

@login_required
def calendario_padres(request):
    tutor = request.user
    hoy = _date.today()

    year = int(request.GET.get("year", hoy.year))
    month = int(request.GET.get("month", hoy.month))

    first_day, total_days = monthrange(year, month)

    # Obtener los niños del tutor (perfil Padre)
    try:
        padre = Padre.objects.get(usuario=request.user)
        ninos = Nino.objects.filter(padre=padre)
    except Padre.DoesNotExist:
        ninos = Nino.objects.none()

    # Planeaciones del mes
    planeaciones = Planeacion.objects.filter(fecha__year=year, fecha__month=month)

    # Novedades de todos los niños del padre (si existen)
    novedades = Novedad.objects.filter(
        nino__in=ninos,
        fecha__year=year,
        fecha__month=month
    ) if ninos.exists() else Novedad.objects.none()

    # Seguimientos de todos los niños del padre (si existen)
    seguimientos = SeguimientoDiario.objects.filter(
        nino__in=ninos,
        fecha__year=year,
        fecha__month=month
    ) if ninos.exists() else SeguimientoDiario.objects.none()

    eventos = {}

    for p in planeaciones:
        dia = p.fecha.day
        eventos.setdefault(dia, {"planeacion": False, "novedad": False})
        eventos[dia]["planeacion"] = True

    for n in novedades:
        dia = n.fecha.day
        eventos.setdefault(dia, {"planeacion": False, "novedad": False})
        eventos[dia]["novedad"] = True

    for s in seguimientos:
        dia = s.fecha.day
        eventos.setdefault(dia, {"planeacion": False, "novedad": False, "seguimiento": False})
        eventos[dia]["seguimiento"] = True

    return render(request, "padre/calendario_padres.html", {
        "year": year,
        "month": month,
        "month_name": MESES.get(month, str(month)),
        "first_day": first_day,
        "total_days": total_days,
        "eventos": eventos,
        "hoy": hoy,
    })


@login_required
def obtener_info(request):
    fecha = request.GET.get("fecha")
    # parsear fecha segura
    fecha_obj = None
    try:
        fecha_obj = _datetime.strptime(fecha, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"planeacion": None, "novedad": None, "seguimientos": []})

    # Obtener los niños desde el perfil del Padre
    try:
        padre = Padre.objects.get(usuario=request.user)
        ninos = Nino.objects.filter(padre=padre)
    except Padre.DoesNotExist:
        ninos = Nino.objects.none()

    planeacion = Planeacion.objects.filter(fecha=fecha_obj).first()
    novedades_del_dia = Novedad.objects.filter(nino__in=ninos, fecha=fecha_obj) if ninos.exists() else Novedad.objects.none()
    
    # --- NUEVA LÓGICA PARA SEGUIMIENTOS ---
    seguimientos_del_dia = SeguimientoDiario.objects.filter(
        nino__in=ninos, fecha=fecha_obj
    ).select_related('nino').prefetch_related('evaluaciones_dimension__dimension') if ninos.exists() else []

    seguimientos_data = []
    for s in seguimientos_del_dia:
        # Construir resumen para padres
        resumen = f"Hoy {s.nino.nombres} se mostró principalmente {s.get_estado_emocional_display().lower()} y su comportamiento fue {s.get_comportamiento_general_display().lower()}."
        if s.observacion_relevante and s.observaciones:
            resumen += f" La madre comunitaria observó: \"{s.observaciones}\"."

        # Obtener evaluaciones por dimensión
        evaluaciones = []
        for ev in s.evaluaciones_dimension.all():
            evaluaciones.append({
                "dimension": ev.dimension.nombre,
                "desempeno": ev.get_desempeno_display()
            })

        seguimientos_data.append({
            "nino_nombre": f"{s.nino.nombres} {s.nino.apellidos}",
            "fecha": s.fecha.strftime("%d/%m/%Y"),
            "resumen_dia": {
                "comportamiento": s.get_comportamiento_general_display(),
                "estado_emocional": s.get_estado_emocional_display(),
                "observacion_relevante": s.observaciones if s.observacion_relevante else None,
                "resumen_para_padres": resumen
            },
            "evaluaciones": evaluaciones,
            "valoracion_dia": s.valoracion,
            "valoracion_restante": 5 - (s.valoracion or 0)
        })

    # --- NUEVA LÓGICA PARA NOVEDADES ---
    novedades_data = []
    for n in novedades_del_dia:
        novedades_data.append({
            "tipo": n.get_tipo_display(),
            "descripcion": n.descripcion,
            "nino_nombre": f"{n.nino.nombres} {n.nino.apellidos}"
        })

    return JsonResponse({
        "planeacion": {
            "nombre": planeacion.nombre_experiencia if planeacion else None,
            "intencionalidad": planeacion.intencionalidad_pedagogica if planeacion else None,
            "materiales": planeacion.materiales_utilizar if planeacion else None
        } if planeacion else None,
        # Se envían como listas para manejar múltiples eventos por día
        "novedades": novedades_data,
        "seguimientos": seguimientos_data,
    })

@login_required
def ver_ficha_nino(request, id):
    nino = get_object_or_404(Nino, id=id)
    return render(request, 'madre/nino_ficha.html', {'nino': nino})

@login_required
def editar_nino(request, id):
    nino = get_object_or_404(Nino, id=id)
    # Asegurarse de que el padre y su usuario existan
    usuario_padre = nino.padre.usuario if nino.padre else None
    perfil_padre = nino.padre if nino.padre else None
    if request.method == 'POST':
        # Se instancian los formularios con los datos del POST, FILES y las instancias de los modelos
        nino_form = NinoForm(request.POST, request.FILES, instance=nino, prefix='nino')
        padre_form = PadreForm(request.POST, request.FILES, instance=usuario_padre, prefix='padre')

        if nino_form.is_valid() and padre_form.is_valid():
            # 💡 VALIDACIÓN: Verificar si el nuevo documento o correo ya existen en otro usuario.
            documento = padre_form.cleaned_data.get('documento')
            correo = padre_form.cleaned_data.get('correo')
            
            if Usuario.objects.filter(Q(documento=documento) | Q(correo=correo)).exclude(id=usuario_padre.id).exists():
                messages.error(request, 'El documento o correo electrónico ingresado ya pertenece a otro usuario.')
            else:
                try:
                    with transaction.atomic():
                        # Guardar el niño con archivos
                        nino_actualizado = nino_form.save(commit=False)
                        # Actualizar campos personalizados
                        nino_actualizado.otro_pais = nino_form.cleaned_data.get('otro_pais', '')
                        # Manejar archivos del niño
                        if 'nino-foto' in request.FILES:
                            nino_actualizado.foto = request.FILES['nino-foto']
                        if 'nino-carnet_vacunacion' in request.FILES:
                            nino_actualizado.carnet_vacunacion = request.FILES['nino-carnet_vacunacion']
                        if 'nino-certificado_eps' in request.FILES:
                            nino_actualizado.certificado_eps = request.FILES['nino-certificado_eps']
                        if 'nino-registro_civil_img' in request.FILES:
                            nino_actualizado.registro_civil_img = request.FILES['nino-registro_civil_img']
                        nino_actualizado.save()
                        
                        # ManyToMany: tipos_discapacidad
                        if nino_actualizado.tiene_discapacidad:
                            nino_actualizado.tipos_discapacidad.set(nino_form.cleaned_data.get('tipos_discapacidad'))
                        else:
                            nino_actualizado.tipos_discapacidad.clear()

                        # Guardar datos del Usuario del padre
                        usuario_actualizado = padre_form.save(commit=False)
                        usuario_actualizado.save()

                        # Actualizar el perfil del padre con archivos
                        if perfil_padre:
                            perfil_padre.ocupacion = padre_form.cleaned_data.get('ocupacion', '')
                            perfil_padre.otra_ocupacion = padre_form.cleaned_data.get('otra_ocupacion', '')
                            perfil_padre.estrato = padre_form.cleaned_data.get('estrato')
                            perfil_padre.telefono_contacto_emergencia = padre_form.cleaned_data.get('telefono_contacto_emergencia', '')
                            perfil_padre.nombre_contacto_emergencia = padre_form.cleaned_data.get('nombre_contacto_emergencia', '')
                            perfil_padre.situacion_economica_hogar = padre_form.cleaned_data.get('situacion_economica_hogar', '')
                            
                            # Manejar archivos del padre
                            if 'padre-documento_identidad_img' in request.FILES:
                                perfil_padre.documento_identidad_img = request.FILES['padre-documento_identidad_img']
                            if 'padre-clasificacion_sisben' in request.FILES:
                                perfil_padre.clasificacion_sisben = request.FILES['padre-clasificacion_sisben']
                            
                            perfil_padre.save()

                        messages.success(request, '¡La información del niño y su tutor ha sido actualizada!')
                        return redirect('listar_ninos')
                except Exception as e:
                    messages.error(request, f"Ocurrió un error inesperado: {e}")

        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')

    else:
        # Se instancian los formularios con los datos existentes para el método GET.
        nino_form = NinoForm(instance=nino, prefix='nino')

        # 💡 CORRECCIÓN: Pre-llenar el formulario del padre con los datos del usuario y del perfil.
        # Los datos del modelo Usuario se cargan con 'instance'.
        # Los datos del modelo Padre (como 'ocupacion') se cargan con 'initial'.
        initial_data_padre = {
            'documento': usuario_padre.documento if usuario_padre else ''
        }
        if perfil_padre:
            initial_data_padre['ocupacion'] = perfil_padre.ocupacion
            initial_data_padre['otra_ocupacion'] = perfil_padre.otra_ocupacion
        
        padre_form = PadreForm(instance=usuario_padre, prefix='padre', initial=initial_data_padre)

    # Obtener el hogar de la madre para el template
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        hogar_madre = nino.hogar  # Usar el hogar del niño como fallback

    return render(request, 'madre/nino_form_nuevo.html', {
        'nino_form': nino_form,
        'padre_form': padre_form,
        'nino': nino, # Se pasa el objeto nino para acceder a datos no editables en la plantilla
        'hogar_madre': hogar_madre,
        'modo_edicion': True,
        'titulo_form': 'Editar Información del Niño'
    })


@login_required
def editar_perfil(request):
    user = request.user
    rol = user.rol.nombre_rol

    # 1. Seleccionar el formulario y la instancia adecuados según el rol
    if rol == 'padre':
        # Para el padre, necesitamos la instancia de su perfil de Padre
        padre_profile, _ = Padre.objects.get_or_create(usuario=user)
        FormClass = PadrePerfilForm
        initial_data = {'ocupacion': padre_profile.ocupacion, 'estrato': padre_profile.estrato}
    elif rol == 'madre_comunitaria':
        FormClass = MadrePerfilForm
        initial_data = None
    else: # administrador
        FormClass = AdminPerfilForm
        initial_data = None

    if request.method == 'POST':
        form = FormClass(request.POST, instance=user, initial=initial_data)
        if form.is_valid():
            # Guardar los datos del modelo Usuario
            user_instance = form.save(commit=False)
            # Si es un padre, guardar los campos adicionales en el modelo Padre
            if rol == 'padre':
                padre_profile.ocupacion = form.cleaned_data.get('ocupacion')
                padre_profile.estrato = form.cleaned_data.get('estrato')
                padre_profile.save()
            user_instance.save()
            messages.success(request, '¡Tu información ha sido actualizada exitosamente!')
            return redirect('editar_perfil')
        else:
            messages.error(request, 'Por favor, corrige los errores en el formulario.')
    else:
        # Al cargar la página, inicializar el formulario con los datos existentes
        form = FormClass(instance=user, initial=initial_data)

    return render(request, 'perfil/editar_perfil.html', {'form': form})

@login_required
def listar_ninos(request):
    # Solo madres comunitarias pueden ver su listado
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Acceso denegado.')
        return redirect('home')
    try:
        hogar = HogarComunitario.objects.get(madre=request.user.madre_profile)
    except HogarComunitario.DoesNotExist:
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')
    
    ninos = Nino.objects.filter(hogar=hogar)
    
    # Contexto con información de matrícula exitosa si existe
    context = {
        'ninos': ninos
    }
    
    # Renderizar el template
    response = render(request, 'madre/nino_list.html', context)
    
    # Limpiar la sesión después de renderizar (para que el mensaje solo se muestre una vez)
    if 'matricula_exitosa' in request.session:
        del request.session['matricula_exitosa']
    if 'cambio_padre_exitoso' in request.session:
        del request.session['cambio_padre_exitoso']
    
    return response

@login_required
def generar_reporte_ninos(request):
    # Aquí puedes generar el PDF o mostrar un mensaje temporal
    return render(request, 'madre/reporte_ninos.html')

@login_required
def eliminar_nino(request, id):
    nino = get_object_or_404(Nino, id=id)
    nino.delete()
    return redirect('listar_ninos')

@login_required
def gestion_ninos(request):
    # 1. Verificar rol y obtener el hogar de la madre
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria': # pragma: no cover
        messages.error(request, 'Acceso denegado.')
        return redirect('home')
    try:
        hogar = HogarComunitario.objects.get(madre=request.user.madre_profile)
    except HogarComunitario.DoesNotExist: # pragma: no cover
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')
    # 2. Filtrar los niños que pertenecen a ese hogar
    ninos_lista = Nino.objects.filter(hogar=hogar).order_by('nombres', 'apellidos')

    # 3. Aplicar paginación
    paginator = Paginator(ninos_lista, 3)  # 3 niños por página, como fue solicitado.
    page_number = request.GET.get('page')
    ninos_paginados = paginator.get_page(page_number)

    return render(request, 'madre/gestion_ninos_list.html', {'ninos': ninos_paginados})

# ----------------------------------------------------
# 💡 NUEVA FUNCIÓN: Cambiar Contraseña del Usuario
# ----------------------------------------------------
@login_required
def cambiar_contrasena(request):
    if request.method == 'POST':
        form = SetPasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Actualizar la sesión para que el usuario no sea deslogueado
            update_session_auth_hash(request, user)
            messages.success(request, '¡Tu contraseña ha sido actualizada exitosamente!')
            # Redirigir a la misma página para que el usuario vea el mensaje de éxito.
            return redirect('cambiar_contrasena')
        else:
            messages.error(request, 'Por favor corrige los errores a continuación.')
    else:
        form = SetPasswordForm(request.user)
    
    return render(request, 'perfil/cambiar_contrasena.html', {'form': form})

@login_required
def detalles_madre_json(request, id):
    try:
        madre_perfil = get_object_or_404(MadreComunitaria, id=id)
        usuario = madre_perfil.usuario
        hogar = HogarComunitario.objects.filter(madre=madre_perfil).first() if hasattr(HogarComunitario, 'madre') else None

        # Obtener la URL de la foto de la madre
        if madre_perfil.foto_madre and hasattr(madre_perfil.foto_madre, 'url'):
            foto_madre_url = madre_perfil.foto_madre.url
        else:
            foto_madre_url = ''

        data = {
            'usuario': {
                'nombres': usuario.nombres,
                'apellidos': usuario.apellidos,
                'tipo_documento': usuario.get_tipo_documento_display(),
                'documento': usuario.documento,
                'correo': usuario.correo,
                'telefono': usuario.telefono,
                'direccion': usuario.direccion,
            },
            'perfil': {
                'nivel_escolaridad': madre_perfil.nivel_escolaridad,
                'titulo_obtenido': madre_perfil.titulo_obtenido or "No especificado",
                'experiencia_previa': madre_perfil.experiencia_previa or "No especificada",
                'foto_madre_url': foto_madre_url,
            },
            'hogar': {
                'nombre_hogar': hogar.nombre_hogar if hogar else "Sin hogar asignado",
                'direccion': hogar.direccion if hogar else "N/A",
                'localidad': hogar.localidad if hogar else "N/A",
                'barrio': hogar.barrio if hogar else "N/A",
                'capacidad_maxima': getattr(hogar, 'capacidad_maxima', 'N/A') if hogar else "N/A",
                'regional': hogar.regional.nombre if hogar and hogar.regional else "N/A",
                'ciudad': hogar.ciudad.nombre if hogar and hogar.ciudad else "N/A",
            }
        }
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def cargar_ciudades(request):

    """
    Retorna JSON con las ciudades de la regional indicada.
    Parámetro: ?regional_id=ID
    """
    regional_id = request.GET.get("regional_id")
    if not regional_id:
        return JsonResponse([], safe=False)
    try:
        ciudades = Ciudad.objects.filter(regional_id=regional_id).order_by('nombre')
        data = [{"id": c.id, "nombre": c.nombre} for c in ciudades]
        return JsonResponse(data, safe=False)
    except Exception:
        return JsonResponse([], safe=False)


# ----------------------------------------------------
# 🆕 NUEVAS VISTAS PARA MEJORAS DE MATRÍCULA
# ----------------------------------------------------

@login_required
def matricular_nino_a_padre_existente(request):
    """Matricular un niño nuevo a un padre que ya existe en el sistema"""
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Solo las madres comunitarias pueden matricular niños.')
        return redirect('home')

    # Obtener el hogar de la madre logueada
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')

    if request.method == 'POST':
        buscar_form = BuscarPadreForm(request.POST, prefix='buscar')
        nino_form = NinoSoloForm(request.POST, request.FILES, prefix='nino')
        
        # Verificar si se ha encontrado un padre
        padre_id = request.POST.get('padre_seleccionado')
        
        if padre_id and nino_form.is_valid():
            try:
                with transaction.atomic():
                    # Obtener el padre seleccionado
                    padre_obj = get_object_or_404(Padre, id=padre_id)
                    
                    # Crear el niño y asociarlo al padre
                    nino = nino_form.save(commit=False)
                    nino.hogar = hogar_madre
                    nino.padre = padre_obj
                    nino.tipo_sangre = nino_form.cleaned_data.get('tipo_sangre')
                    nino.parentesco = nino_form.cleaned_data.get('parentesco')
                    nino.tiene_discapacidad = nino_form.cleaned_data.get('tiene_discapacidad', False)
                    nino.otra_discapacidad = nino_form.cleaned_data.get('otra_discapacidad', '')
                    nino.otro_pais = nino_form.cleaned_data.get('otro_pais', '')
                    nino.save()
                    
                    # ManyToMany: tipos_discapacidad
                    if nino.tiene_discapacidad:
                        nino.tipos_discapacidad.set(nino_form.cleaned_data.get('tipos_discapacidad'))
                    else:
                        nino.tipos_discapacidad.clear()
                    
                    messages.success(request, f'Niño {nino.nombres} matriculado correctamente al padre {padre_obj.usuario.get_full_name()}.')
                    return redirect('listar_ninos')
                    
            except Exception as e:
                messages.error(request, f"Ocurrió un error inesperado: {e}")
        else:
            if not padre_id:
                messages.error(request, 'Debe buscar y seleccionar un padre antes de continuar.')
            if not nino_form.is_valid():
                messages.error(request, 'Por favor corrige los errores en el formulario del niño.')
    else:
        buscar_form = BuscarPadreForm(prefix='buscar')
        nino_form = NinoSoloForm(prefix='nino')

    return render(request, 'madre/matricular_a_padre_existente.html', {
        'hogar_madre': hogar_madre,
        'buscar_form': buscar_form,
        'nino_form': nino_form,
    })


@login_required
def buscar_padre_ajax(request):
    """Vista AJAX para buscar padre por documento en matriculación a padre existente"""
    documento = request.GET.get('documento', '').strip()
    data = {'encontrado': False, 'padre': None}

    if documento:
        try:
            # Buscar usuario con rol de padre
            usuario_padre = Usuario.objects.filter(
                documento=documento, 
                rol__nombre_rol='padre',
                is_active=True
            ).first()
            
            if usuario_padre:
                try:
                    padre_profile = Padre.objects.get(usuario=usuario_padre)
                    data.update({
                        'encontrado': True,
                        'padre': {
                            'id': padre_profile.id,
                            'nombres': usuario_padre.nombres,
                            'apellidos': usuario_padre.apellidos,
                            'documento': usuario_padre.documento,
                            'correo': usuario_padre.correo,
                            'telefono': usuario_padre.telefono,
                            'direccion': usuario_padre.direccion,
                            'ocupacion': padre_profile.get_ocupacion_display() if padre_profile.ocupacion else 'No especificada'
                        }
                    })
                except Padre.DoesNotExist:
                    data['mensaje'] = 'Usuario encontrado pero sin perfil de padre completo.'
            else:
                data['mensaje'] = 'No se encontró un padre con ese documento.'
        except Exception as e:
            data['error'] = f'Error en la búsqueda: {str(e)}'

    return JsonResponse(data)


@login_required
def cambiar_padre_de_nino(request):
    """Cambiar la asignación de padre de un niño existente"""
    if not hasattr(request.user, 'rol') or request.user.rol.nombre_rol != 'madre_comunitaria':
        messages.error(request, 'Solo las madres comunitarias pueden cambiar asignaciones.')
        return redirect('home')

    # Obtener el hogar de la madre logueada
    try:
        madre_profile = request.user.madre_profile
        hogar_madre = HogarComunitario.objects.get(madre=madre_profile)
    except (MadreComunitaria.DoesNotExist, HogarComunitario.DoesNotExist):
        messages.error(request, 'No tienes un hogar comunitario asignado.')
        return redirect('madre_dashboard')

    if request.method == 'POST':
        cambiar_form = CambiarPadreForm(hogar=hogar_madre, data=request.POST, prefix='cambiar')
        buscar_form = BuscarPadreForm(request.POST, prefix='buscar')
        padre_form = PadreForm(request.POST, request.FILES, prefix='padre')
        
        # Verificar qué acción se está realizando
        accion = request.POST.get('accion')
        nino_id = request.POST.get('nino_seleccionado')
        nuevo_padre_id = request.POST.get('nuevo_padre_id')
        motivo_cambio = request.POST.get('motivo_cambio')
        observaciones_motivo = request.POST.get('observaciones_motivo', '')
        
        # Log para debugging
        print(f"[DEBUG] Acción: {accion}")
        print(f"[DEBUG] Niño ID: {nino_id}")
        print(f"[DEBUG] Nuevo Padre ID: {nuevo_padre_id}")
        print(f"[DEBUG] Motivo: {motivo_cambio}")
        print(f"[DEBUG] Observaciones: {observaciones_motivo}")
        print(f"[DEBUG] POST data: {request.POST}")
        
        if accion == 'cambiar_existente' and nino_id and nuevo_padre_id:
            # Validar que se haya seleccionado un motivo
            if not motivo_cambio:
                messages.error(request, 'Debe seleccionar el motivo del cambio de asignación.')
                return redirect('cambiar_padre_nino')
            
            # Cambiar a padre existente
            print(f"[DEBUG] Cambiando niño {nino_id} a padre existente {nuevo_padre_id}")
            try:
                with transaction.atomic():
                    nino = get_object_or_404(Nino, id=nino_id, hogar=hogar_madre)
                    nuevo_padre = get_object_or_404(Padre, id=nuevo_padre_id)
                    padre_anterior = nino.padre
                    
                    print(f"[DEBUG] Niño encontrado: {nino.nombres} {nino.apellidos}")
                    print(f"[DEBUG] Padre anterior: {padre_anterior.usuario.get_full_name() if padre_anterior else 'Sin padre'}")
                    print(f"[DEBUG] Nuevo padre: {nuevo_padre.usuario.get_full_name()}")
                    
                    nino.padre = nuevo_padre
                    nino.save()
                    
                    mensaje_exito = f'El niño {nino.nombres} {nino.apellidos} ha sido asignado correctamente '
                    if padre_anterior:
                        mensaje_exito += f'del padre {padre_anterior.usuario.get_full_name()} '
                    mensaje_exito += f'al padre {nuevo_padre.usuario.get_full_name()}.'
                    
                    # Guardar en la sesión para el SweetAlert
                    request.session['cambio_padre_exitoso'] = {
                        'nombre': f'{nino.nombres} {nino.apellidos}',
                        'mensaje': mensaje_exito,
                        'motivo': motivo_cambio,
                        'observaciones': observaciones_motivo
                    }
                    
                    messages.success(request, mensaje_exito)
                    print(f"[DEBUG] Cambio exitoso - Motivo: {motivo_cambio}")
                    return redirect('listar_ninos')
            except Exception as e:
                print(f"[ERROR] Error al cambiar padre: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error al cambiar padre: {e}")
                
        elif accion == 'cambiar_nuevo' and nino_id and padre_form.is_valid():
            # Validar que se haya seleccionado un motivo
            if not motivo_cambio:
                messages.error(request, 'Debe seleccionar el motivo del cambio de asignación.')
                return redirect('cambiar_padre_nino')
            
            # Cambiar a nuevo padre
            print(f"[DEBUG] Creando nuevo padre y cambiando niño {nino_id}")
            try:
                with transaction.atomic():
                    nino = get_object_or_404(Nino, id=nino_id, hogar=hogar_madre)
                    padre_anterior = nino.padre
                    
                    # Crear nuevo usuario padre
                    rol_padre = Rol.objects.get(nombre_rol='padre')
                    doc_padre = padre_form.cleaned_data['documento']
                    tipo_doc = padre_form.cleaned_data['tipo_documento']
                    
                    usuario_padre, created = Usuario.objects.get_or_create(
                        documento=doc_padre,
                        rol=rol_padre,
                        defaults={
                            'tipo_documento': tipo_doc,
                            'nombres': padre_form.cleaned_data['nombres'],
                            'apellidos': padre_form.cleaned_data['apellidos'],
                            'correo': padre_form.cleaned_data['correo'],
                            'telefono': padre_form.cleaned_data['telefono'],
                            'direccion': padre_form.cleaned_data.get('direccion', ''),
                        }
                    )

                    if created:
                        usuario_padre.set_password(str(doc_padre))
                        usuario_padre.save()

                    # Crear perfil del padre
                    padre_obj, _ = Padre.objects.get_or_create(usuario=usuario_padre)
                    padre_obj.ocupacion = padre_form.cleaned_data.get('ocupacion', '')
                    padre_obj.otra_ocupacion = padre_form.cleaned_data.get('otra_ocupacion', '')
                    padre_obj.estrato = padre_form.cleaned_data.get('estrato')
                    padre_obj.telefono_contacto_emergencia = padre_form.cleaned_data.get('telefono_contacto_emergencia', '')
                    padre_obj.nombre_contacto_emergencia = padre_form.cleaned_data.get('nombre_contacto_emergencia', '')
                    padre_obj.situacion_economica_hogar = padre_form.cleaned_data.get('situacion_economica_hogar', '')
                    padre_obj.save()
                    
                    # Actualizar niño
                    nino.padre = padre_obj
                    nino.save()
                    
                    messages.success(request, 
                        f'El niño {nino.nombres} {nino.apellidos} ha sido asignado correctamente '
                        f'del padre {padre_anterior.usuario.get_full_name()} '
                        f'al nuevo padre {padre_obj.usuario.get_full_name()}.'
                    )
                    return redirect('listar_ninos')
            except Exception as e:
                print(f"[ERROR] Error al crear nuevo padre y cambiar: {str(e)}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Error al cambiar padre: {e}")
        else:
            print(f"[DEBUG] Datos incompletos o formulario inválido")
            print(f"[DEBUG] Acción válida: {accion in ['cambiar_existente', 'cambiar_nuevo']}")
            print(f"[DEBUG] Niño ID presente: {bool(nino_id)}")
            if accion == 'cambiar_existente':
                print(f"[DEBUG] Nuevo padre ID presente: {bool(nuevo_padre_id)}")
            if accion == 'cambiar_nuevo':
                print(f"[DEBUG] Padre form válido: {padre_form.is_valid()}")
                if not padre_form.is_valid():
                    print(f"[DEBUG] Errores del formulario: {padre_form.errors}")
            messages.error(request, 'Datos incompletos para realizar el cambio.')
    else:
        cambiar_form = CambiarPadreForm(hogar=hogar_madre, prefix='cambiar')
        buscar_form = BuscarPadreForm(prefix='buscar')
        padre_form = PadreForm(prefix='padre')

    return render(request, 'madre/cambiar_padre_nino.html', {
        'hogar_madre': hogar_madre,
        'cambiar_form': cambiar_form,
        'buscar_form': buscar_form,
        'padre_form': padre_form,
    })

@login_required
def actualizar_foto_perfil(request):
    """
    Vista AJAX para que la madre comunitaria actualice su foto de perfil
    directamente desde el navbar.
    """
    if request.method == 'POST' and request.user.rol.nombre_rol == 'madre_comunitaria':
        try:
            madre_profile = get_object_or_404(MadreComunitaria, usuario=request.user)
            
            # El nombre del input en el form debe ser 'foto_madre'
            new_photo = request.FILES.get('foto_madre')

            if new_photo:
                madre_profile.foto_madre = new_photo
                madre_profile.save()
                
                # Devolver la URL de la nueva foto
                return JsonResponse({'success': True, 'new_photo_url': madre_profile.foto_madre.url})
            else:
                return JsonResponse({'success': False, 'error': 'No se recibió ningún archivo.'}, status=400)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)

    return JsonResponse({'success': False, 'error': 'Método no permitido o rol incorrecto.'}, status=405)


@login_required
def subir_documentos_nino(request):
    """
    Vista AJAX para subir documentos faltantes de un niño desde el dashboard de la madre.
    """
    if request.method == 'POST':
        try:
            # Verificar que sea madre comunitaria
            if request.user.rol.nombre_rol != 'madre_comunitaria':
                return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción.'}, status=403)
            
            # Obtener el hogar de la madre
            hogar_madre = HogarComunitario.objects.filter(madre=request.user.madre_profile).first()
            if not hogar_madre:
                return JsonResponse({'success': False, 'error': 'No tienes un hogar asignado.'}, status=400)
            
            # Obtener el ID del niño
            nino_id = request.POST.get('nino_id')
            if not nino_id:
                return JsonResponse({'success': False, 'error': 'No se especificó el niño.'}, status=400)
            
            # Verificar que el niño pertenezca al hogar de la madre
            nino = get_object_or_404(Nino, id=nino_id, hogar=hogar_madre)
            
            # Actualizar documentos
            documentos_actualizados = []
            
            # Registro civil
            if 'registro_civil_img' in request.FILES:
                nino.registro_civil_img = request.FILES['registro_civil_img']
                documentos_actualizados.append('Registro Civil')
            
            # Carnet de vacunación
            if 'carnet_vacunacion' in request.FILES:
                nino.carnet_vacunacion = request.FILES['carnet_vacunacion']
                documentos_actualizados.append('Carnet de Vacunación')
            
            # Certificado EPS
            if 'certificado_eps' in request.FILES:
                nino.certificado_eps = request.FILES['certificado_eps']
                documentos_actualizados.append('Certificado EPS')
            
            # Foto
            if 'foto' in request.FILES:
                nino.foto = request.FILES['foto']
                documentos_actualizados.append('Foto')
            
            # Guardar cambios
            if documentos_actualizados:
                nino.save()
                return JsonResponse({
                    'success': True,
                    'message': f'Documentos actualizados: {", ".join(documentos_actualizados)}',
                    'documentos_actualizados': documentos_actualizados
                })
            else:
                return JsonResponse({'success': False, 'error': 'No se seleccionó ningún documento para subir.'}, status=400)
            
        except Nino.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'El niño no existe o no pertenece a tu hogar.'}, status=404)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': f'Error al procesar los documentos: {str(e)}'}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
